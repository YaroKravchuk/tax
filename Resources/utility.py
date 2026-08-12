import os
import pandas as pd
import platform
import subprocess
from collections import namedtuple
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# The book of record for every load. It lives outside this repo, two folders up, and
# nothing in the program works without it
BOOK_RECORDS_FILE = '../../Dump Trucking BookRecords - TEST.xlsx'

# Sheets holding load data are named for their year, e.g. "2026 Dump Trucking"
YEAR_SHEET_MARKER = 'Dump Trucking'

# A blank PROJECT ID cell comes back from astype(str) as the text "nan", so every
# spelling of "no project here" has to be skipped when listing a sheet's projects
BLANK_PROJECT_IDS = {'', 'nan', 'nat', 'none'}

# Columns the project picker reads. The fuller list a finished run needs is checked
# later by validate_data, once the data has been narrowed to one project
PICKER_COLUMNS = ['PROJECT ID', 'DATE', 'CUSTOMER']

# One project as the picker shows it. Two folded-down copies of its text are kept, both
# made once up front so that filtering thousands of projects on every keystroke stays
# cheap. search_text is the ID by itself, which is what names one project exactly, and is
# what the box holds once one has been picked. match_text is the ID and the customer
# together, which is what the search reads, so a job can be found by whoever it was for
Project = namedtuple('Project', 'id customer loads first_date last_date search_text match_text')

# Everything a run needs to write its workbooks
Materials = namedtuple('Materials', 'driver_log_wb driver_log_template invoice_wb invoice_sheet '
                                    'data min_date max_date')

# The driver log's table of loads: the row its headings sit on, and the last row a load
# can be written to
LOAD_TABLE_HEADING_ROW = 8
LOAD_TABLE_LAST_ROW = 18

# The columns that table is laid out in, as the heading, the first and last spreadsheet
# column it covers, and the column its look is borrowed from. A truck's time in and out
# are kept once for its whole day, in the Trucking boxes under the table, so the table
# carries no time of its own and the width the two time columns held is shared out among
# the columns that are left. Only the merges move, and only onto edges of columns the
# template already has, so the table still ends where it did and every other block on the
# sheet keeps the size and the place it was drawn with.
#
# populate_driver_log_sheet writes each load into the first column of the column it
# belongs in, so these letters and the ones it uses have to say the same thing
LOAD_TABLE_COLUMNS = (
    ('Loaded From', 'A', 'A', None),
    ('Delivered To', 'B', 'D', None),
    ('Material', 'E', 'H', None),
    ('Qty', 'I', 'J', 'H'),
    ('Mat.$', 'K', 'K', None),
    ('DumpFee', 'L', 'N', None),
    ('SB Time', 'O', 'P', 'P'),
)


class BookRecords:
    """The BookRecords workbook, read once and then kept.

    The form needs the load data to offer projects to pick from, and the generator
    needs the same data to fill the sheets. Reading it once here means a big workbook
    is parsed a single time per run instead of once for each of them.
    """

    def __init__(self, path=BOOK_RECORDS_FILE):
        self.path = path
        self._sheets = {}
        self._projects = {}
        try:
            # Noted before anything is read, so that what the form reports about the file
            # is the state of the file the numbers actually came from
            self.full_path = os.path.abspath(path)
            self.read_mtime = os.path.getmtime(path)
            self.last_saved = datetime.fromtimestamp(self.read_mtime)
            # Closed the moment its sheet names have been read. Left to itself pandas
            # holds the file open until the garbage collector happens to come round to
            # it, which here meant holding it for as long as the form was up: Windows
            # will not let Excel save over a file another program has open, and refuses
            # the save as a sharing violation. Saving the records while the form is open
            # is the very thing the line at the top of the form is there to report on,
            # so it has to be something the records can survive
            with pd.ExcelFile(path) as workbook:
                self.sheet_names = workbook.sheet_names
        except FileNotFoundError:
            raise ValueError('The load records could not be found.'
                             f'\n\nThe program expected them here:\n{os.path.abspath(path)}')

    # Function to spot the records being saved again while the form is still open, which
    # would leave the form offering data that is already out of date
    def has_been_saved_since_read(self):
        try:
            return os.path.getmtime(self.path) != self.read_mtime
        except OSError:
            return False

    # Function to list the year sheets that hold load data, newest year first
    def year_sheet_names(self):
        return sorted((name for name in self.sheet_names if YEAR_SHEET_MARKER in name), reverse=True)

    # Function to pick the year sheet to start on, which is this year's when there is one
    def default_year_sheet(self):
        sheet_names = self.year_sheet_names()
        current_year = str(datetime.now().year)
        return next((name for name in sheet_names if current_year in name), sheet_names[0])

    # Function to read one year sheet, holding on to it so a second look is free
    def sheet(self, sheet_name):
        if sheet_name not in self._sheets:
            df = pd.read_excel(self.path, sheet_name=sheet_name)

            missing_columns = [column for column in PICKER_COLUMNS if column not in df.columns]
            if missing_columns:
                raise ValueError(f'The sheet "{sheet_name}" is missing the '
                                 f'{", ".join(missing_columns)} column(s), so its loads cannot be read.')

            # Project IDs are compared as text everywhere, so they are squared away once here
            df['PROJECT ID'] = df['PROJECT ID'].astype(str).str.strip()
            self._sheets[sheet_name] = df

        return self._sheets[sheet_name]

    # Function to list every project on a year sheet, the most hauled first, each with the
    # customer, load count and dates that let the right one be recognised
    def projects(self, sheet_name):
        if sheet_name not in self._projects:
            self._projects[sheet_name] = self._build_project_list(sheet_name)
        return self._projects[sheet_name]

    def _build_project_list(self, sheet_name):
        df = self.sheet(sheet_name)
        named = df[~df['PROJECT ID'].str.lower().isin(BLANK_PROJECT_IDS)]
        if named.empty:
            return []

        summary = named.groupby('PROJECT ID', sort=False).agg(
            customer=('CUSTOMER', 'last'),
            loads=('PROJECT ID', 'size'),
            first_date=('DATE', 'min'),
            last_date=('DATE', 'max'),
        )

        projects = []
        for project_id, row in summary.iterrows():
            # A project with no customer recorded is still searchable by its address
            searchable = project_id if pd.isna(row.customer) else f'{project_id} {row.customer}'
            projects.append(Project(project_id, row.customer, int(row.loads), row.first_date,
                                    row.last_date, project_id.lower(), searchable.lower()))

        # The biggest jobs first, then addresses in alphabetical order so that the many
        # projects sharing a load count keep a settled place in the list rather than
        # shuffling about with whatever order the sheet happens to hold them in. The
        # already folded-down address is sorted on, so case never decides the order
        projects.sort(key=lambda project: (-project.loads, project.search_text))
        return projects


# Function to put a project's loads in the order the sheets are built from. A driver log
# sheet is started whenever the date or the truck changes from the load before, so every
# load for one truck on one day has to arrive in one run. A load recorded out of order -
# a day's hauling typed in after the next day's had been - would otherwise open a second
# sheet for a truck that already has one, splitting the day across two tickets that both
# count their loads from one, while the invoice listed the same day twice.
#
# Dates and trucks keep the order the records introduce them in rather than being sorted
# into an order of their own, so gathering the loads is all this does. Sorting on the
# truck would reorder the tickets as well, and it would order them by their spelling:
# "PR TR #11" comes before "PR TR #4" when the two are read as text
def order_loads(data):
    ordered = data.sort_values('DATE', kind='stable')
    groups = ordered.groupby(['DATE', 'TRUCK ID#'], sort=False, dropna=False).ngroup()
    return ordered.loc[groups.sort_values(kind='stable').index]


# Function to lay the table of loads out again, in the columns LOAD_TABLE_COLUMNS names.
# The template still carries a Time In and a Time Out column against every load, from
# before the times were kept once for the whole day, and this is what takes them off it
def rebuild_load_table(sheet):
    rows = range(LOAD_TABLE_HEADING_ROW, LOAD_TABLE_LAST_ROW + 1)

    for row in rows:
        for merged in [str(area) for area in sheet.merged_cells.ranges if area.min_row == row]:
            sheet.unmerge_cells(merged)

    # Two of the columns now begin at a cell the template only ever used as the middle of
    # another one, which carries no border, fill or heading of its own, so each borrows
    # the look of the column whose place it takes. That is done before anything is merged
    # again: merging replaces the cells it swallows, and one of the columns being
    # borrowed from is swallowed by the column in front of it
    borrowed_looks = {}
    for heading, first, last, borrow_from in LOAD_TABLE_COLUMNS:
        if borrow_from is not None:
            for row in rows:
                borrowed_looks[first, row] = copy(sheet[f'{borrow_from}{row}']._style)

    for heading, first, last, borrow_from in LOAD_TABLE_COLUMNS:
        for row in rows:
            cell = sheet[f'{first}{row}']
            if borrow_from is not None:
                cell._style = borrowed_looks[first, row]
            # The DumpFee column starts where Time In used to, and its cells are still
            # formatted as times. Nothing in the table holds a time any more
            cell.number_format = 'General'
            if first != last:
                sheet.merge_cells(f'{first}{row}:{last}{row}')

        # Merging clears the headings of the columns that have been swallowed, and the
        # ones left standing are named again, since most of them have moved along
        sheet[f'{first}{LOAD_TABLE_HEADING_ROW}'] = heading


# Function to narrow a year sheet down to one project inside a date range. The form and
# the generator both call this, so what the form promises is exactly what gets built
def filter_loads(records, sheet_name, project_id, start_date, end_date):
    data = records.sheet(sheet_name)
    data = data[data['PROJECT ID'] == project_id]
    if pd.notna(start_date):
        data = data[data['DATE'] >= start_date]
    if pd.notna(end_date):
        data = data[data['DATE'] <= end_date]
    return order_loads(data)


# Function to open the folder holding the finished files, so they do not have to be hunted for
def open_output_folder(folder):
    try:
        if platform.system() == 'Windows':
            os.startfile(folder)
        else:
            subprocess.run(['open', folder], check=False)
    except Exception as e:
        print(f"Could not open the folder {folder}: {str(e)}")


# Function to create new materials like workbooks, the data table, and template sheets
def create_materials(choices):
    driver_log_wb = load_workbook(filename='MASTER_DumpTruck_TimeSheet_ProspectLLC_2025_FINAL.xlsx')
    driver_log_template = driver_log_wb["2024 Version"]
    driver_log_template['B2'].font = Font(name='Calibri', color='FFFFFF', size=18, b=True)

    # Done before the headings are dressed below, so that the cells named there are the
    # ones the table is actually headed with
    rebuild_load_table(driver_log_template)

    bold_cells = "G1 K1 A6 F6 N6 A8 B8 E8 I8 K8 L8 O8 A16 A17 A18 A20 A21 F16 J16 J17 J18 J20 J21".split()
    for cell in bold_cells:
        driver_log_template[cell].font = Font(name='Calibri', color='FFFFFF', size=11.5, b=True)

    data = filter_loads(choices.records, choices.sheet_name, choices.project_id,
                        choices.start_date, choices.end_date)
    validate_data(data, choices.project_id)

    invoice_template = ('InvoiceASAP_Template_2025.xlsx' if choices.taxable
                        else 'InvoiceASAP_Template_2025_NonTaxable.xlsx')
    invoice_wb = load_workbook(filename=invoice_template)
    invoice_wb["Blank_Template"].title = "Invoice"
    invoice_sheet = invoice_wb["Invoice"]
    min_date = data['DATE'].min()
    max_date = data['DATE'].max()
    invoice_sheet["D5"] = choices.project_id
    invoice_sheet["G2"] = f"Start: {min_date.strftime('%m/%d/%y')}"
    invoice_sheet["H2"] = f"End: {max_date.strftime('%m/%d/%y')}"

    # Delete template sheet from final workbook file
    del driver_log_wb['2024 Version']

    return Materials(driver_log_wb, driver_log_template, invoice_wb, invoice_sheet, data,
                     min_date.strftime('%b %d').upper(), max_date.strftime('%b %d').upper())


# Function to check if BookRecords data exists and contains the correct columns
def validate_data(data, project_id):
    if data.empty:
        raise ValueError("No data found for project ID " + project_id)

    required_columns = ["PROJECT ID", "DATE", "TRUCK ID#", "PRODUCT", "LOAD QTY \n",
                        "CUSTOMER", "HAULING FROM", "HAULING TO"]
    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        raise ValueError(f"Required data columns are missing: {', '.join(missing_columns)}")
