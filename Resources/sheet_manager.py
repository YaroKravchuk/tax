from openpyxl.drawing.image import Image
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
import math
import os
import pandas as pd
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles.colors import BLUE
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.units import pixels_to_EMU as p2e

# Folder holding the logo and the signature images
IMAGE_FOLDER = 'pngs'

# Cells on the driver log that each image sits in
LOGO_CELLS = 'O1:P4'
APPROVER_SIGNATURE_CELLS = 'M25:P25'
DRIVER_SIGNATURE_CELLS = 'B25:D25'

# Gap left between an image and the edge of the cells holding it, in points, so ink
# never touches the black border around the cell
IMAGE_PADDING = 2

# The logo is a brand mark, so it keeps a fixed size in pixels rather than filling its cells
LOGO_WIDTH = 90
LOGO_HEIGHT = 65

# Excel measures column widths in "characters", meaning the width of a digit in the
# workbook's default font. That font is Times New Roman 10 here, whose digits are 5
# points wide, and no extra padding is added on top. Both numbers were measured against
# Excel itself, so re-measure them if the template's default font is ever changed.
POINTS_PER_CHARACTER = 5.0
DEFAULT_COLUMN_CHARACTERS = 9.0
DEFAULT_ROW_POINTS = 15

# Images are sized in pixels, while cells are measured in points. There are 96 pixels
# and 72 points to an inch, and a point is 12700 English Metric Units.
POINTS_PER_PIXEL = 0.75
EMU_PER_POINT = 12700


# Function to find the signature image file for a driver. Signature files live in the pngs
# folder and are named after the driver's name in BookRecords, e.g. "Vlad" -> Vlad_signature.png
def get_driver_signature_file(driver_name):
    if not pd.notna(driver_name):
        return None

    signature_file = os.path.join(IMAGE_FOLDER, str(driver_name).strip().capitalize() + "_signature.png")
    if not os.path.exists(signature_file):
        print(f"No signature image found for driver '{driver_name}' (looked for {signature_file})")
        return None

    return signature_file


# Function to shrink an image so it fits inside the given box while keeping its aspect ratio
def scale_image_to_fit(image, max_width, max_height):
    scale = min(max_width / image.width, max_height / image.height, 1)
    image.width = round(image.width * scale)
    image.height = round(image.height * scale)


# Function to measure one column in points. Excel lays columns out on whole points,
# and rounding here reproduces every column width it reports for this template.
def column_points(sheet, column_index):
    column = sheet.column_dimensions.get(get_column_letter(column_index))
    characters = column.width if column is not None and column.width else DEFAULT_COLUMN_CHARACTERS
    return round(characters * POINTS_PER_CHARACTER)


# Function to measure one row in points. Excel drops the fraction of a row height
# rather than rounding it, which is why row 25 measures 32 points here and not 32.5.
def row_points(sheet, row_index):
    row = sheet.row_dimensions.get(row_index)
    return math.floor(row.height if row is not None and row.height else DEFAULT_ROW_POINTS)


# Function to measure a block of cells in points, which is the unit Excel positions in
def cells_size(sheet, cell_range):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    width = sum(column_points(sheet, index) for index in range(min_col, max_col + 1))
    height = sum(row_points(sheet, index) for index in range(min_row, max_row + 1))
    return width, height


# Function to turn a distance from the top left of a block of cells into an anchor.
# An image is pinned to one cell plus an offset, and Excel ignores any part of that
# offset reaching past the cell it belongs to, so the offset has to be carried across
# into later columns and rows until what is left fits inside a single cell. LibreOffice
# is forgiving about this, so getting it wrong looks fine in the PDF but not in Excel.
def cell_anchor(sheet, cell_range, left_offset, top_offset):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    column_index = min_col
    while column_index < max_col and left_offset >= column_points(sheet, column_index):
        left_offset -= column_points(sheet, column_index)
        column_index += 1

    row_index = min_row
    while row_index < max_row and top_offset >= row_points(sheet, row_index):
        top_offset -= row_points(sheet, row_index)
        row_index += 1

    return AnchorMarker(col=column_index - 1, colOff=round(left_offset * EMU_PER_POINT),
                        row=row_index - 1, rowOff=round(top_offset * EMU_PER_POINT))


# Function to place an image in the middle of a block of cells. Measuring the cells and
# working the position out from them keeps every image centred by construction, so the
# template's rows and columns can be resized without images drifting out of place
def add_image_to_cells(sheet, image_path, cell_range, width=None, height=None):
    image = Image(image_path)
    area_width, area_height = cells_size(sheet, cell_range)

    if width is not None and height is not None:
        image.width, image.height = width, height
    else:
        scale_image_to_fit(image,
                           (area_width - IMAGE_PADDING * 2) / POINTS_PER_PIXEL,
                           (area_height - IMAGE_PADDING * 2) / POINTS_PER_PIXEL)

    image_width = image.width * POINTS_PER_PIXEL
    image_height = image.height * POINTS_PER_PIXEL

    marker = cell_anchor(sheet, cell_range,
                         (area_width - image_width) / 2, (area_height - image_height) / 2)
    image.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(p2e(image.width), p2e(image.height)))
    sheet.add_image(image)


class SheetManager:

    # Function to initialize SheetManager class
    def __init__(self, driver_log_wb, driver_log_template, invoice_sheet, taxable):

        # Initialize resources
        self.driver_log_wb = driver_log_wb
        self.driver_log_template = driver_log_template
        self.invoice_sheet = invoice_sheet
        self.taxable = taxable

        # Create global variables
        self.load_row_count = 0
        self.row_count = 0
        self.sheet_count = 0
        self.prev_date = None
        self.prev_truck = None
        self.driver_log_sheet = None
        self.curr_date_row = "8"

    # Creates a new driver log sheet
    def create_new_driver_log_sheet(self, row):
        self.sheet_count = self.sheet_count + 1
        self.load_row_count = 0
        self.prev_date = row["DATE"]
        self.prev_truck = row["TRUCK ID#"]
        self.driver_log_sheet = self.driver_log_wb.copy_worksheet(self.driver_log_template)
        self.driver_log_sheet.title = "Driver Log " + str(self.sheet_count)

        self.driver_log_sheet['A4'] = CellRichText(
            'P.O. Box 3571 Bellevue, WA 98009'
            '\n(425) 569-9910 / ',
            TextBlock(InlineFont(rFont='Calibri', b=True, sz=9), 'prospectllc@pauldimov.com'),
        )
        self.driver_log_sheet['G4'] = CellRichText(
            TextBlock(InlineFont(rFont='Calibri', i=True, sz=7),
                      'And whatever you do, do it heartily, as to the Lord and not to men, knowing that from the Lord '
                      'you will receive the reward of the inheritance; for you serve the Lord Christ.',),
            TextBlock(InlineFont(rFont='Calibri', b=True, sz=7), '\nColossians 3:23-24 NKJV'),
        )

        self.driver_log_sheet['H1'] = row["DATE"]
        self.driver_log_sheet['A7'] = row["CUSTOMER"]
        self.driver_log_sheet['F7'] = row["PROJECT ID"]
        self.driver_log_sheet['N7'] = row["TRUCK ID#"]

        driver_name = row.get("DRIVER'S NAME")
        if pd.notna(driver_name):
            self.driver_log_sheet['B24'] = driver_name

        self.add_images_to_driver_log(driver_name)

    # Function to input a single row of driver log data
    def populate_driver_log_sheet(self, row):
        # First check if previous sheet is same truck, same day
        if row["DATE"] == self.prev_date and row["TRUCK ID#"] == self.prev_truck:
            self.load_row_count = self.load_row_count + 1
        else:
            # If data row contains a new truck or a new day, create a new sheet
            self.create_new_driver_log_sheet(row)

        # The columns are the ones LOAD_TABLE_COLUMNS lays the table out in. A load is
        # written with no time of its own: a truck's time in and out cover its whole day
        # rather than any one load, and are kept once, in the Trucking boxes below
        driver_log_row = str(self.load_row_count + 9)
        self.driver_log_sheet['A' + driver_log_row] = row["HAULING FROM"]
        self.driver_log_sheet['B' + driver_log_row] = row["HAULING TO"]
        self.driver_log_sheet['E' + driver_log_row] = row["PRODUCT"]
        self.driver_log_sheet['I' + driver_log_row] = row["LOAD QTY \n"]
        if pd.notna(row["MATERIAL COST"]):
            self.driver_log_sheet['K' + driver_log_row] = "X"
        if pd.notna(row["DUMP FEE RATE"]):
            self.driver_log_sheet['L' + driver_log_row] = "X"
        self.driver_log_sheet['O' + driver_log_row] = row["STAND-BY TIME"]

        if not pd.isnull(row["TIME IN"]):
            self.driver_log_sheet['M20'] = row["TIME IN"]
        if not pd.isnull(row["TIME OUT"]):
            self.driver_log_sheet['M21'] = row["TIME OUT"]
        if not pd.isnull(row["NOTES"]):
            # Store current "comment" text in current_value and add a new line (If the cell has any text in it)
            current_value = str(self.driver_log_sheet['F21'].value) + "\n" \
                if (self.driver_log_sheet['F21'].value is not None
                    and str(self.driver_log_sheet['F21'].value).strip() != "") \
                else ""
            # Format comment: Load 1: "COMMENT"
            self.driver_log_sheet['F21'] = (current_value + "Load " + str(self.load_row_count + 1) + ': "'
                                            + str(row["NOTES"]) + '"')
        if row["HOURS"] > 0:
            self.driver_log_sheet['M22'] = row["HOURS"]

    # Function to input a single row of invoice data
    def populate_invoice_sheet_row(self, row):

        if self.row_count + 8 > 67:
            # raise ValueError("There is too much data to fit on the invoice! \n\nDecrease time range...")
            self.row_count = 999
            return

        invoice_row = str(self.row_count + 8)
        self.set_cell_value("D4", row.get("CUSTOMER"))
        self.set_cell_value(f"A{invoice_row}", row.get("DATE").date() if pd.notna(row.get("DATE")) else None)
        self.set_cell_value(f"B{invoice_row}", row.get("TRUCK ID#"))
        self.set_cell_rich_text(f"C{invoice_row}", row.get("SERVICE TYPE"), row.get("PRODUCT"))
        self.set_cell_value(f"E{invoice_row}", row.get("LOAD QTY \n"))
        self.set_cell_value(f"F{invoice_row}", row.get("RATE PER LOAD"))

        if self.taxable and pd.notna(row.get("RATE PER LOAD")):
            self.invoice_sheet["G" + invoice_row] = "X"

        if pd.notna(row.get("LOAD QTY \n")) and pd.notna(row.get("RATE PER LOAD")):
            self.set_cell_value(f"H{invoice_row}", f"=E{invoice_row}*F{invoice_row}")

        self.row_count = self.row_count + 1

        offset = 27
        if pd.notna(row.get("STAND-BY TIME")) and pd.notna(row.get("STAND-BY RATE")):
            self.populate_invoice_sheet_row_subcategory(
                row,
                f"{' ' * offset}↳ Truck Standby Hours",
                True,
                row["STAND-BY TIME"],
                row["STAND-BY RATE"]
            )
        if pd.notna(row["TIME IN"]):
            self.populate_invoice_sheet_row_subcategory(
                row,
                f"{' ' * offset}↳ Truck Hours Worked",
                True,
                row["HOURS"],
                row["RATE PER HOUR"]
            )
        if pd.notna(row["DUMP FEE RATE"]):
            self.populate_invoice_sheet_row_subcategory(
                row,
                f"{' ' * offset}↳ Dump Fee",
                False,
                row["LOAD QTY \n"],
                row["DUMP FEE RATE"]
            )
        if pd.notna(row["MATERIAL COST"]):
            self.populate_invoice_sheet_row_subcategory(
                row,
                f"{' ' * offset}↳ Material Cost",
                False,
                row["LOAD QTY \n"],
                row["MATERIAL COST"]
            )

    def populate_invoice_sheet_row_subcategory(self, row, description, is_hours_unit, unit, rate):
        if not pd.notna(unit) or not pd.notna(rate):
            return

        if self.row_count + 8 > 67:
            # raise ValueError("There is too much data to fit on the invoice! \n\nDecrease time range...")
            self.row_count = 999
            return

        invoice_row = str(self.row_count + 8)

        self.set_cell_value(f"A{invoice_row}", row.get("DATE").date() if pd.notna(row.get("DATE")) else None)
        self.set_cell_value(f"B{invoice_row}", row.get("TRUCK ID#"))
        self.set_cell_value(f"C{invoice_row}", description.ljust(30))

        if is_hours_unit:
            self.invoice_sheet[f"E{invoice_row}"].number_format = '0.00'

        self.set_cell_value(f"E{invoice_row}", unit)
        self.set_cell_value(f"F{invoice_row}", rate)

        if self.taxable and is_hours_unit:
            self.invoice_sheet["G" + invoice_row] = "X"
        
        self.set_cell_value(f"H{invoice_row}", f"=E{invoice_row}*F{invoice_row}")

        self.row_count = self.row_count + 1

    def set_cell_value(self, cell_reference, value):
        try:
            if pd.notna(value):
                self.invoice_sheet[cell_reference] = value
        except Exception as e:
            print(f"Error processing subcategory row {self.row_count + 8}: {str(e)}")
            print(f"cell_reference: {cell_reference}, value: {value}")

    def set_cell_rich_text(self, cell_reference, service_type, product):
        if pd.notna(service_type) and pd.notna(product):
            self.invoice_sheet[cell_reference] = CellRichText(
                TextBlock(InlineFont(rFont='Tahoma', color=BLUE, sz=9, b=True),
                          f"{' ' * 20}{service_type}:"),
                TextBlock(InlineFont(rFont='Tahoma', color=BLUE, sz=9),
                          f" {product}")
            )

    # Function to configure the Logo image and the Signature images for the driver log
    def add_images_to_driver_log(self, driver_name):
        add_image_to_cells(self.driver_log_sheet, os.path.join(IMAGE_FOLDER, 'logo.png'), LOGO_CELLS,
                           width=LOGO_WIDTH, height=LOGO_HEIGHT)
        add_image_to_cells(self.driver_log_sheet, os.path.join(IMAGE_FOLDER, 'sig.png'),
                           APPROVER_SIGNATURE_CELLS)

        self.add_driver_signature_to_driver_log(driver_name)

    # Function to add the driver's signature image under "Driver's Name" / "Signature"
    def add_driver_signature_to_driver_log(self, driver_name):
        signature_file = get_driver_signature_file(driver_name)
        if signature_file is None:
            return

        add_image_to_cells(self.driver_log_sheet, signature_file, DRIVER_SIGNATURE_CELLS)

    def merge_date_cells(self):
        current_val = None
        start_row = 8

        for row in range(8, self.row_count + 8):
            cell_value = self.invoice_sheet[f"A{row}"].value

            if cell_value != current_val:
                if current_val and start_row < row - 1:
                    self.invoice_sheet.merge_cells(f"A{start_row}:A{row-1}")
                current_val = cell_value
                start_row = row

            if current_val and row == self.row_count + 7 and start_row < self.row_count + 7:
                self.invoice_sheet.merge_cells(f"A{start_row}:A{self.row_count + 7}")

    def merge_truck_cells(self):
        start_row = 8

        for row in range(8, self.row_count + 8):
            cell_value = self.invoice_sheet[f"C{row}"].value

            if '↳' not in str(cell_value):
                if start_row < row - 1:
                    self.invoice_sheet.merge_cells(f"B{start_row}:B{row-1}")
                start_row = row
            elif row == self.row_count + 7:
                self.invoice_sheet.merge_cells(f"B{start_row}:B{self.row_count + 7}")
