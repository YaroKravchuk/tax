import pandas as pd
import PySimpleGUI as sg
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl import load_workbook
from openpyxl.styles import Font


# Function to create the layout of the UI
def create_layout():
    sg.ChangeLookAndFeel('GreenTan')
    form = sg.FlexForm('Audit', default_element_size=(40, 1))
    # Open the Excel file
    excel_file = pd.ExcelFile('../../Dump Trucking BookRecords - TEST.xlsx')
    # Get all sheet names
    sheet_names = [name for name in excel_file.sheet_names if 'Dump Trucking' in name]
    # Check if current year exists in sheet names, and if so, set it as default
    current_year = str(datetime.now().year)
    default_sheet = next((name for name in sheet_names if current_year in name), None)
    layout = [
        [sg.Text('Generate Audit Daily Load Tickets!', size=(30, 1), font=("Helvetica", 25))],
        [sg.Text('"Dump Trucking" Year Sheet')],
        [sg.Combo(sheet_names, size=(30, 1), default_value=default_sheet)],
        [sg.Text('Project ID')],
        [sg.InputText()],
        [sg.Text('Start Date')],
        [sg.InputText()],
        [sg.Text('End Date')],
        [sg.InputText()],
        [sg.Text('_' * 80)],
        [sg.Submit(), sg.Cancel()]
    ]
    return form, layout


# Function to configure the Logo image and the Signature image for the driver log
def create_images():
    p2e = pixels_to_EMU
    logoImage = Image('img.jpg')
    signatureImage = Image('sig.jpg')

    logoImage_X_Coordinate = 755
    logoImage_Y_Coordinate = 11
    logoImage_Width = 98
    logoImage_Height = 101
    logoImage.anchor = AbsoluteAnchor(pos=XDRPoint2D(p2e(logoImage_X_Coordinate), p2e(logoImage_Y_Coordinate)),
                                      ext=XDRPositiveSize2D(p2e(logoImage_Width), p2e(logoImage_Height)))

    signatureImage_X_Coordinate = 755
    signatureImage_Y_Coordinate = 477
    signatureImage_Width = 104
    signatureImage_Height = 40
    signatureImage.anchor = AbsoluteAnchor(pos=XDRPoint2D(p2e(signatureImage_X_Coordinate), p2e(signatureImage_Y_Coordinate)),
                                           ext=XDRPositiveSize2D(p2e(signatureImage_Width), p2e(signatureImage_Height)))

    return logoImage, signatureImage


# Function to format the Driver Log Excel sheet
def create_workbook():
    wb = load_workbook(filename='../../Template.xlsx')
    source = wb["Sheet1"]
    source['B2'].font = Font(color='FFFFFF', size=18, b=True)
    source['B5'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['E5'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['B7'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['D7'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['H7'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['O7'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['B17'].font = Font(color='FFFFFF', size=11.5, b=True)
    source['I19'].font = Font(color='FFFFFF', size=11.5, b=True)

    ws = wb.create_sheet('Data')
    ws.append(["Date", "Truck ID", "Product", "QTY"])
    ws.column_dimensions['A'].width = 13
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 7

    return wb, ws, source


def validate_data(data):
    if data.empty:
        raise ValueError("No data found for the provided project ID.")

    required_columns = ["PROJECT ID", "DATE", "TRUCK ID#", "PRODUCT", "LOAD QTY \n", "CUSTOMER", "HAULING FROM", "HAULING TO"]
    missing_columns = [col for col in required_columns if col not in data.columns]
    if missing_columns:
        raise ValueError(f"Required data columns are missing: {', '.join(missing_columns)}")


def populate_data_sheet(ws, row_count, row):
    ws["A" + str(row_count)] = row["DATE"].date()
    ws["B" + str(row_count)] = row["TRUCK ID#"]
    ws["C" + str(row_count)] = row["PRODUCT"]
    ws["D" + str(row_count)] = row["LOAD QTY \n"]


def populate_driver_log_sheet(target, load_row_count, row):
    target['E' + str(load_row_count + 6)] = row["PROJECT ID"]
    target['B' + str(load_row_count + 8)] = row["HAULING FROM"]
    target['D' + str(load_row_count + 8)] = row["HAULING TO"]
    target['H' + str(load_row_count + 8)] = row["PRODUCT"]
    target['L' + str(load_row_count + 8)] = row["LOAD QTY \n"]